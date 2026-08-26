from pathlib import Path
import json
from openpyxl import load_workbook

source = Path('/home/ubuntu/upload/Agilite.xlsx')
out = Path('/home/ubuntu/agilite-collection-tracker/data/catalog.json')
out.parent.mkdir(parents=True, exist_ok=True)
wb = load_workbook(source, data_only=True)
ws = wb['Master']
headers = {str(ws.cell(1, c).value or '').strip(): c for c in range(1, ws.max_column + 1)}

def value(row, key):
    col = headers.get(key)
    return ws.cell(row, col).value if col else None

def clean(v):
    if v is None:
        return ''
    return str(v).strip()

records = []
for row in range(2, ws.max_row + 1):
    name = clean(value(row, 'Product Name'))
    if not name:
        continue
    records.append({
        'id': clean(value(row, 'ID')) or str(len(records) + 1),
        'name': name,
        'category': clean(value(row, 'Category')) or 'Collectible',
        'variant': clean(value(row, 'Variant/Options')),
        'sku': clean(value(row, 'SKU(s)')),
        'usUrl': clean(value(row, 'US URL')),
        'israelUrl': clean(value(row, 'Israel URL')),
        'internationalUrl': clean(value(row, 'International URL')),
        'usAvailability': clean(value(row, 'US availability')),
        'israelAvailability': clean(value(row, 'Israel availability')),
        'internationalAvailability': clean(value(row, 'International availability')),
        'source': clean(value(row, 'Source')),
        'evidenceClass': clean(value(row, 'Evidence class')),
        'notes': clean(value(row, 'Notes')),
    })

out.write_text(json.dumps(records, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
print(f'Wrote {len(records)} records to {out}')
